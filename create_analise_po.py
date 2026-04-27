from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ABA 1
ws1 = wb.active
ws1.title = "Analise Detalhada"

headers = ["Chamado", "Titulo", "Modulo", "Resumo da Necessidade", "Item Relacionado", "Tipo", "Acao", "Relevancia"]

DARK_BLUE = "1F3864"
LIGHT_BLUE = "D9E2F3"
WHITE = "FFFFFF"

thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

header_font = Font(name="Arial", bold=True, color=WHITE, size=10)
header_fill = PatternFill("solid", start_color=DARK_BLUE)
body_font = Font(name="Arial", size=9)

for col, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

data = [
    ("DSI-33742","Separar campos CATMAT no cadastro do produto","Almoxarifado / CATMAT","O cliente quer que os campos CATMAT (BNDASAF e e-SUS) sejam separados no cadastro do produto, evitando a necessidade de acessar o modulo de almoxarifado para ajustes. Reduz trabalho repetitivo e melhora a manutencao dos dados.","SAUD-15837","Caracteristica","Relacionar","Media"),
    ("DSI-33566","Corrigindo Codigo e-SUS no CATMAT","Exportacao e-SUS / CATMAT","Ao clicar para corrigir inconsistencia de CATMAT pelo exportador e-SUS, o sistema abre o cadastro do produto que nao e o local correto. O cliente precisa de um facilitador direto no exportador para corrigir o campo sem navegar pelo almoxarifado.","SAUD-40828","Caracteristica","Relacionar","Alta"),
    ("DSI-33648","Bases e-SUS AF - BNAFAR e RNDS","Assistencia Farmaceutica / Exportacoes","Com o lancamento do e-SUS AF e descontinuacao do Horus, o cliente solicita verificacao se o sistema precisa de adequacoes para integracao com BNAFAR e RNDS. Alerta estrategico sobre nova plataforma nacional de assistencia farmaceutica.","SAUD-62481","Caracteristica","Relacionar","Alta"),
    ("PROIEPP-19622","Integracao eSUS-AF (substituicao do Horus)","Assistencia Farmaceutica / Exportacoes","O sistema Horus sera descontinuado em ate 180 dias. Municipios precisam que o sistema suporte integracao com o novo eSUS-AF (HUB-AF/MedSUS) para nao haver ruptura na gestao da assistencia farmaceutica.","SAUD-62481","Caracteristica","Relacionar","Alta"),
    ("PUBINFOS-56527","Pre-visualizacao de medicamentos na dispensacao","Farmacia / Dispensacao","Na tela de dispensacao, o usuario precisa abrir cada registro individualmente para ver os medicamentos. O cliente solicita uma aba de pre-visualizacao que exiba os medicamentos sem precisar abrir cada dispensacao.","SAUD-59960","Story","Relacionar","Alta"),
    ("BTHSC-307489","Na dispensa permitir adicionar quantidade a maior","Farmacia / Dispensacao","O sistema nao permite dispensar quantidade superior ao prescrito, mesmo para medicamentos nao controlados. Na pratica, medico prescreve em comprimidos mas a dispensacao ocorre por cartela, exigindo flexibilidade.","--","--","Criar Story","--"),
    ("BTHSC-307305","Relatorio de dispensacao diaria portaria 344","Farmacia / Relatorios","O cliente precisa de relatorio diario simplificado de dispensacao de medicamentos controlados e antibioticos com data, paciente, medicamento, quantidade e prescritor - sem as capas do relatorio atual.","SAUD-69439","Story","Relacionar","Alta"),
    ("PUBINFOS-56398","Agendamento de grupos e replicar em agenda por ordem de chegada","Agendamento","Fisioterapeutas atendem grupos de pacientes por consulta e precisam replicar agendamentos em agendas por ordem de chegada. O sistema nao permite nem agendamento em grupo nem replicacao nesse tipo de agenda.","--","--","Criar Story","--"),
    ("CSMCSM-148314","Configurar DDD das ligacoes de IA","Voicebot / Confirmacao de Agendamentos","O DDD das ligacoes automaticas de confirmacao por IA gera desconfianca nos pacientes (suspeitam de golpe). O cliente precisa configurar o DDD local nas ligacoes para aumentar a taxa de atendimento.","SAUD-28801","Caracteristica","Relacionar","Alta"),
    ("BTHSC-305243","Afericao de pressao: ordenar do mais recente para o mais antigo","Atendimento / Triagem","Na triagem, os registros de afericao de pressao aparecem em ordem crescente (mais antigo primeiro), dificultando a visualizacao da ultima afericao. O cliente precisa que a ordenacao padrao seja da data mais recente.","--","--","Criar Story","--"),
    ("BTHSC-305114","Filtro por equipe (INE) no cofinanciamento","Cofinanciamento","Unidades com mais de uma equipe nao conseguem analisar o grafico de evolucao do cofinanciamento por equipe individualmente. O gestor precisa visualizar indicadores separados por equipe e por unidade.","SAUD-73717","Story","Relacionar","Alta"),
    ("SOL-4874","Bloquear impressao de resultado de procedimento sem assinatura","Resultados / Permissoes","Botao de impressora na aba RES > Procedimentos imprime laudo sem assinatura digital e nao ha permissao para revogar esse acesso. O cliente precisa controlar quem pode imprimir resultados nao assinados.","--","--","Criar Story","--"),
    ("BW-28416","Melhoria na tela de APAC da Central de Regulacao","Regulacao / APAC","Na tela de APAC da regulacao, o cliente quer dar continuidade as solicitacoes com o mesmo fluxo de passos das telas de procedimentos e consultas (aguardando, agendado, com data). Hoje a APAC nao permite esse acompanhamento.","--","--","Criar Story","--"),
    ("SOL-4865","Assinatura eletronica com validacao em exames laboratoriais","Laboratorio / Assinatura Digital","A RDC 978/25 exige assinatura eletronica com validacao nos resultados de exames no momento da liberacao, nao apenas na impressao. Atualmente o sistema acopla a assinatura so ao imprimir, e apenas de quem imprime.","SAUD-59962","Story","Relacionar","Alta"),
    ("SOL-4863","Permissao de recepcionar via agendamento bloqueavel por unidade","Agendamento / Permissoes","Profissionais com acesso a agendamento em outras unidades conseguem recepcionar e registrar ausencias onde nao tem permissao, podendo manipular vagas. A permissao precisa ser revogavel e validada pela unidade do agendamento.","--","--","Criar Story","--"),
    ("DSI-33703","Adicionar procedimento puericultura automaticamente","Atendimento / Puericultura","O sistema deve adicionar o procedimento 03.01.01.027-7 automaticamente ao avaliar fatores de risco na puericultura, como o e-SUS ja faz. Evita esquecimento e facilita o cumprimento do indicador.","SAUD-73453","Story","Relacionar","Media"),
    ("CSMCSM-148215","Mais de um local/horario de desembarque no transporte","Transporte / Viagem","Pacientes com mais de um compromisso no mesmo dia de viagem precisam que o sistema suporte multiplos locais e horarios de desembarque para que o motorista possa se programar.","SAUD-46090","Story","Relacionar","Alta"),
    ("DSI-33447","Desobrigar campos adicionais em mamografia e citopatologico","Atendimento / Requisicao de Exames","Ao lancar procedimentos de mamografia e coleta citopatologica, o sistema exige campos adicionais que o e-SUS nao exige, tornando o processo mais lento. O cliente pede que esses campos sejam opcionais.","SAUD-69389","Story","Relacionar","Alta"),
    ("CSMCSM-147602","Centros de custo obrigatorios na requisicao externa","Almoxarifado / Requisicoes","Na requisicao externa do almoxarifado, os campos de centro de custo sao opcionais mas deveriam ser obrigatorios, como ja ocorre na requisicao interna. Diferentes profissionais gerenciam diferentes tipos de produtos por centro de custo.","SAUD-29550","Story","Relacionar","Alta"),
    ("SOL-4848","Vacina dose unica com aplicacao anual nao deveria exigir justificativa","Vacinacao","A vacina Influenza Trivalente e dose unica mas tem aplicacao anual. Toda vez que e aplicada, o sistema exige justificativa de replicacao, o que nao faz sentido para a rotina dos vacinadores.","--","--","Criar Story","--"),
    ("CSMCSM-147259","Alerta de exame repetido no agendamento de procedimento","Agendamento / Procedimentos","Ao agendar um procedimento, o sistema deve alertar se o paciente ja realizou o mesmo exame nos ultimos X dias configurados. O municipio usa laboratorio sem atendimento medico e precisa do alerta no agendamento.","--","--","Criar Story","--"),
    ("DSI-33468","Informacoes aba Recursos do Municipio com valores estimados","Cofinanciamento","O cliente aguarda melhoria que exiba as notas finais com valores estimados a receber na aba Recursos do Municipio, conforme nota tecnica. Atualmente a visao do gestor e limitada sem os valores corretos.","SAUD-73717","Story","Relacionar","Alta"),
    ("CSMCSM-147174","Interfaceamento de aparelhos de Hematologia e Bioquimica","Laboratorio / Equipamentos","O laboratorio municipal precisa que os aparelhos de hematologia e bioquimica sejam integrados ao sistema para obtencao automatica dos resultados, eliminando a digitacao manual e otimizando o trabalho diario.","SAUD-8351","Caracteristica","Relacionar","Alta"),
    ("SOL-4836","Transferencia de clientes em viagem nao gera registro rastreavel","Transporte / Transferencia","Ao transferir pacientes entre viagens, o agendamento original nao muda de situacao e nenhum registro da nova viagem e gerado, tornando impossivel rastrear o percurso do paciente.","--","--","Criar Story","--"),
    ("DSI-33435","Visualizacao do agente restrita a sua equipe e microarea","Domiciliar e Territorial / Permissoes","No modulo domiciliar e territorial, agentes conseguem ver registros de outros profissionais. O cliente pede que a permissao possa ser restringida para exibir apenas registros da propria equipe e microarea.","--","--","Criar Story","--"),
    ("DSI-33430","Solicitacao de exames/procedimentos OCI","Atendimento / Requisicoes","O cliente pede que o sistema permita registrar requisicoes de Oferta de Cuidado Integrado (OCI), seguindo o modelo do e-SUS APS/PEC. Necessario para conformidade com o fluxo de atencao primaria.","--","--","Criar Story","--"),
    ("DSI-33424","Inativar imoveis no modulo territorial","Domiciliar e Territorial / Imoveis","No modulo territorial, nao e possivel inativar imoveis demolidos, abandonados ou fora da area de abrangencia. O cliente precisa dessa funcionalidade para gerenciar territorios com precisao.","--","--","Criar Story","--"),
    ("CSMCSM-146969","Saldo real de cotas disponivel visivel sem hover","Regulacao / Cotas","Na programacao orcamentaria, o sistema exibe o valor total inicial de cotas. O cliente precisa ver o saldo real disponivel diretamente nas colunas, sem precisar passar o mouse sobre o campo de consumo.","SAUD-21807","Story","Relacionar","Alta"),
    ("BTHSC-296613","Campo observacao por item da requisicao (nao apenas global)","Atendimento / Requisicao de Exames","No e-SUS, cada item de uma requisicao tem sua propria justificativa. No sistema Betha, o campo de observacao e unico por requisicao global. O cliente precisa de observacao individual por item para rastreabilidade.","SAUD-46132","Story","Relacionar","Alta"),
    ("BTHSC-295799","Botao Selecionar Todos na transferencia de agendamentos","Agendamento / Transferencia","Ao transferir agendamentos de um profissional para outro, e necessario selecionar cada agendamento individualmente. O cliente pede um botao Selecionar Todos para transferencias em massa por ausencia do profissional.","SAUD-23316","Story","Relacionar","Media"),
    ("SCCT-2053","Programa MedCasa na dispensacao","Farmacia / Dispensacao / Domiciliar","O municipio precisa incluir o Programa MedCasa no modulo de dispensa de medicamentos e/ou na saude domiciliar, para registrar e controlar dispensacoes vinculadas a esse programa especifico.","--","--","Criar Story","--"),
    ("BTHSC-291218","Observacoes nao saem na impressao das medicacoes","Atendimento / Impressao","As observacoes adicionadas as medicacoes para administracao na unidade nao aparecem na impressao gerada para a enfermagem. O profissional perde informacoes importantes no documento impresso.","SAUD-68155","Story","Relacionar","Media"),
    ("BTHSC-290833","Erro na transferencia: vinculacao incorreta com NF de origem","Farmacia / Transferencia entre Unidades","Ao gerar etiquetas de uma transferencia, o sistema busca dados da NF original da unidade remetente em vez da transferencia atual. Gera quantidade incorreta, travamento e impossibilidade de edicao.","--","--","Criar Story","--"),
    ("CSMCSM-145346","Historico de consumo 30/60 dias na tela de requisicao","Almoxarifado / Requisicoes","Na tela de requisicao, o sistema exibe apenas o estoque atual. O farmaceutico precisa ver o consumo dos ultimos 30 e 60 dias para embasar pedidos assertivos e evitar excesso ou falta de estoque.","--","--","Criar Story","--"),
    ("CSMCSM-144694","Bloquear impressao de documentos por perfil","Atendimento / Permissoes","Recepcionistas com acesso ao menu de atendimentos conseguem imprimir documentos da consulta medica, violando a privacidade do paciente. E necessaria permissao especifica para bloquear a impressao por perfil.","--","--","Criar Story","--"),
    ("BTHSC-277244","Observacoes de Enfermagem nao imprimem na ficha de atendimento","Atendimento / Triagem / Impressao","As informacoes do campo Observacoes de Enfermagem durante a triagem aparecem no RES mas nao sao impressas na Ficha de Atendimento (FA/FAA). A enfermagem precisa dessas informacoes no documento fisico.","SAUD-68155","Story","Relacionar","Media"),
    ("BTHSC-275821","Botao Hoje na agenda","Agendamento / Usabilidade","O sistema vicia em uma data ao navegar pela agenda, exigindo varios cliques para voltar ao dia atual. O cliente pede um botao fixo Hoje para retornar rapidamente ao dia corrente.","--","Em avaliacao","Verificar internamente","--"),
    ("CSMCSM-143238","Vinculacao de beneficio com saldo de estoque","Beneficios / Almoxarifado","Nutricionistas iniciam solicitacoes de beneficios sem saber se ha estoque disponivel. O cliente pede que o sistema mostre apenas os suplementos com saldo disponivel ao solicitar um beneficio.","--","--","Criar Story","--"),
    ("BTHSC-271287","Baixa de medicamento sem cadastrar paciente (Horus)","Farmacia / Dispensacao","O sistema Horus permite saida de medicamento para usuario SUS sem cadastrar o paciente quando nao ha documentos completos. O cliente precisa de funcionalidade equivalente no Saude Betha.","--","--","Criar Story","--"),
    ("BTHSC-271131","Entrada sem lote/validade nao registra fabricante","Almoxarifado / Entradas","Ao dar entrada de itens sem lote e validade, o sistema gera um lote generico e nao registra o fabricante informado na entrada, perdendo rastreabilidade importante.","--","--","Criar Story","--"),
    ("DSI-32602","Preencher CID automaticamente quando ha apenas um obrigatorio","Atendimento / SIGTAP","Quando um procedimento tem apenas um CID obrigatorio no SIGTAP, o sistema deveria preenchê-lo automaticamente. Hoje o profissional precisa informar manualmente, aumentando o tempo de atendimento.","--","--","Criar Story","--"),
    ("SOL-4458","Campo Intervencoes/Procedimentos Clinicos com CIAP2","Atendimento / CIAP2","No e-SUS PEC, o profissional registra intervencoes usando CIAP2 e SIGTAP. O sistema Betha so possui SIGTAP. O psicologo precisa registrar codigo CIAP2 que nao existe na tabela SIGTAP.","--","--","Criar Story","--"),
    ("STAFSIS-261739","Profissional - Usuario do Sistema na Secretaria ve todos os usuarios","Cadastros / Profissionais","No cadastro de profissionais, ao selecionar Secretaria como local, o campo Usuario do Sistema deveria exibir todos os usuarios da entidade. Atualmente exibe apenas os da unidade, impedindo vinculacao de gestores.","SAUD-46060","Story","Relacionar","Media"),
    ("DSI-32179","Rotina de reprocessamento avancado pos-ajuste cadastral","Exportacoes / e-SUS / BPA","Apos ajustes em cadastros, as fichas geradas anteriormente nao sao reprocessadas. Solicita-se rotina administrativa com agendamento noturno para reprocessar fichas de exportacao automaticamente.","--","--","Criar Story","--"),
    ("BTHSC-252170","Alterar veiculo em viagem criada pela agenda","Transporte / Viagens","Viagens criadas a partir da agenda bloqueiam a alteracao do veiculo. Na pratica, a demanda de pacientes muda ao longo do dia, exigindo troca de veiculo. Sem isso e necessario criar nova viagem e redistribuir todos os pacientes.","SAUD-72679","Story","Relacionar","Alta"),
    ("SOL-4431","Procedimento com agenda especifica ainda agendavel na agenda geral","Agendamento / Agenda Especifica","Quando existe agenda especifica para um procedimento em determinada unidade, o sistema ainda permite agenda-lo na agenda geral da mesma unidade, sem opcao de bloqueio.","--","--","Criar Story","--"),
    ("STAFSIS-261236","Desbloqueio de acesso bloqueado por tentativas de senha","Gerenciador de Acessos","Quando o acesso de um usuario e bloqueado por tentativas incorretas de senha, o administrador da Secretaria nao consegue fazer o desbloqueio - precisa acionar suporte tecnico.","--","--","Criar Story","--"),
    ("STAFSIS-261234","Alteracao de senha pelo administrador","Gerenciador de Acessos","O administrador da Secretaria nao consegue redefinir a senha de profissionais diretamente no sistema. Precisa depender de suporte tecnico para isso, atrasando o atendimento.","--","--","Criar Story","--"),
    ("STAFSIS-261233","Agendamento em massa por profissional","Agendamento / Central de Regulacao","A Central de Regulacao precisa agendar multiplos pacientes para um mesmo profissional sem repetir o fluxo individual para cada um. O cliente propoe interface para selecionar varios pacientes e horarios de uma vez.","--","--","Criar Story","--"),
    ("SOL-4349","Confirmar comparecimento/falta em agendamentos externos","Agendamento Externo / Regulacao","O municipio paga convenios com base na confirmacao de comparecimento do paciente ao agendamento externo. Atualmente o agendamento externo ja e finalizado imediatamente, sem possibilidade de registrar presenca/falta posterior.","SAUD-45483","Story","Relacionar","Alta"),
    ("BTHSC-217586","Vagas de viagem TFD separadas por ida e volta","Transporte / TFD","O atributo de vagas no TFD e consolidado sem distinguir sentido da viagem (ida/volta). Alguns pacientes ficam na cidade de destino e o controle de vagas precisa refletir essa realidade por sentido.","--","--","Criar Story","--"),
    ("SOL-4107","Falha de seguranca: PDFs acessiveis via URL sem autenticacao","Seguranca / Documentos","URLs de PDFs gerados pelo sistema (laudos, resultados de exames) sao publicas e acessiveis por qualquer pessoa na internet sem autenticacao, expondo dados sensiveis de pacientes.","--","--","Criar Story","--"),
    ("SOL-4044","QR code de validacao junto a assinatura digital nos laudos","Laboratorio / Assinatura Digital","Para conformidade com RDC 978/25 e Lei 14.063/2020, o QR code de validacao deve ser impresso na mesma pagina da assinatura digital escaneada do profissional liberador - nao em pagina separada.","SAUD-58774","Story","Relacionar","Media"),
    ("SOL-4043","Recepcionar coleta de exames terceirizados na mesma unidade","Laboratorio / Terceirizacao","O laboratorio municipal coleta amostras de exames que serao enviados a terceirizados junto com exames proprios. O sistema nao suporta esse fluxo, exigindo workaround de recepcionar e depois cancelar o atendimento.","--","--","Criar Story","--"),
]

fill_even = PatternFill("solid", start_color=LIGHT_BLUE)
fill_odd = PatternFill("solid", start_color=WHITE)
colors_acao = {
    "Relacionar": "C6EFCE",
    "Criar Story": "FFEB9C",
    "Verificar internamente": "DDDDDD",
}

for i, row in enumerate(data, 2):
    fill = fill_even if i % 2 == 0 else fill_odd
    for col, val in enumerate(row, 1):
        c = ws1.cell(row=i, column=col, value=val)
        c.font = body_font
        c.fill = fill
        c.border = border
        if col == 4:
            c.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            c.alignment = Alignment(vertical="top", wrap_text=False)
        if col == 7:
            ac = colors_acao.get(val)
            if ac:
                c.fill = PatternFill("solid", start_color=ac)

col_widths = [16, 45, 28, 70, 16, 16, 22, 12]
for col, width in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(col)].width = width
for row in range(2, len(data) + 2):
    ws1.row_dimensions[row].height = 60
ws1.row_dimensions[1].height = 30
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"

# ABA 2
ws2 = wb.create_sheet("Resumo Executivo")
headers2 = ["Acao", "Quantidade", "Chamados"]
DARK_GREEN = "1E4620"

for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = PatternFill("solid", start_color=DARK_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

resumo = [
    ("Relacionar a Story", 18,
     "PUBINFOS-56527; BTHSC-307305; BTHSC-305114; SOL-4865; DSI-33703; CSMCSM-148215; DSI-33447; CSMCSM-147602; DSI-33468; CSMCSM-146969; BTHSC-296613; BTHSC-295799; BTHSC-291218; BTHSC-277244; STAFSIS-261739; BTHSC-252170; SOL-4349; SOL-4044"),
    ("Relacionar a Caracteristica", 7,
     "DSI-33742; DSI-33566; DSI-33648; PROIEPP-19622; CSMCSM-148314; CSMCSM-147174; BTHSC-307489"),
    ("Criar nova Story", 28,
     "BTHSC-307489; PUBINFOS-56398; BTHSC-305243; SOL-4874; BW-28416; SOL-4863; SOL-4848; CSMCSM-147259; SOL-4836; DSI-33435; DSI-33430; DSI-33424; SCCT-2053; BTHSC-290833; CSMCSM-145346; CSMCSM-144694; CSMCSM-143238; BTHSC-271287; BTHSC-271131; DSI-32602; SOL-4458; DSI-32179; SOL-4431; STAFSIS-261236; STAFSIS-261234; STAFSIS-261233; BTHSC-217586; SOL-4107; SOL-4043"),
    ("Verificar internamente", 1, "BTHSC-275821"),
    ("TOTAL", "=SUM(B2:B4)", "54 chamados analisados em 23/04/2026"),
]

row_colors2 = {
    "Relacionar a Story": "C6EFCE",
    "Relacionar a Caracteristica": "BDD7EE",
    "Criar nova Story": "FFEB9C",
    "Verificar internamente": "DDDDDD",
    "TOTAL": "F2F2F2",
}

for i, row in enumerate(resumo, 2):
    bg = row_colors2.get(row[0], WHITE)
    for col, val in enumerate(row, 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.fill = PatternFill("solid", start_color=bg)
        c.border = border
        c.font = Font(name="Arial", bold=(row[0] == "TOTAL"), size=9)
        c.alignment = Alignment(horizontal="center" if col == 2 else "left", vertical="top", wrap_text=True)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 100
for row in range(1, 7):
    ws2.row_dimensions[row].height = 40
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = "A1:C1"

# Legenda
ws2.cell(row=8, column=1, value="Legenda").font = Font(name="Arial", bold=True, size=9)
legend = [
    ("Relacionar a Story", "C6EFCE"),
    ("Relacionar a Caracteristica", "BDD7EE"),
    ("Criar nova Story", "FFEB9C"),
    ("Verificar internamente", "DDDDDD"),
]
for i, (label, color) in enumerate(legend, 9):
    c = ws2.cell(row=i, column=1, value=label)
    c.fill = PatternFill("solid", start_color=color)
    c.font = Font(name="Arial", size=9)
    c.border = border
    c.alignment = Alignment(vertical="center")
    ws2.row_dimensions[i].height = 18

OUTPUT = "C:/Claude/Skill/analise-po-melhorias.xlsx"
wb.save(OUTPUT)
print(f"OK: {OUTPUT}")
